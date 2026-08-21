"""Excel .xlsx text extractor (openpyxl). Deterministic for a fixed lib version.

A spreadsheet is not an opaque source — it is a zip of XML with a fully
specified data layer, and parsing it is a faithful transform (the Core-boundary
test). Before this extractor existed, an .xlsx fell to the bytes-only path and
adapters reached for `model-read`: inference standing in for data that was
sitting right there (T-234, filed from a live failure).

The aid is **per-sheet markdown tables** — sheet titles as headings, the first
row rendered as the table header (the common case; when row 1 is data it is
still shown, nothing is lost). Cached values, not formulas (`data_only=True`):
what the sheet last displayed is what the aid records, and the choice is
stamped in the aid's footer. Big sheets truncate at a fixed row cap with an
in-band marker — surface, never silently — and the canonical bytes always hold
the rest (ADR-0010 rule 1).

Registered by the registry only when `openpyxl` imports, so a box without it
captures .xlsx bytes-only (ADR-0010 rule 5) rather than crashing. The old
binary `.xls` format is a different beast and is intentionally not handled
here, same as `.doc` in the docx extractor.
"""
from __future__ import annotations

import io

import openpyxl as _openpyxl

from .base import Extractor

try:
    from importlib.metadata import version as _pkg_version
    _VERSION = _pkg_version("openpyxl")
except Exception:  # pragma: no cover
    _VERSION = "?"

#: Per-sheet row cap. Past this the aid records that it truncated, in-band.
MAX_ROWS_PER_SHEET = 500


def _cell(value) -> str:
    if value is None:
        return ""
    # Pipes would break the table row; newlines would break the table line.
    return str(value).replace("|", "\\|").replace("\n", " ").replace("\r", " ")


def _sheet_facts(ws) -> str:
    """One line of strictly mechanical facts about a sheet (T-048): declared
    dimensions plus typed min/max over ALL cells - a full lazy walk, because a
    range computed over a truncated prefix would be a wrong fact, and the whole
    point is that these facts stay trustworthy where the aid's table truncates.
    Deliberately interpretation-free: no header guessing, no column semantics -
    "which row is the header" is judgment, and judgment is summary territory
    (the schema READING is authored at derive time, per the ingest flow)."""
    import datetime as _dt
    n_rows = ws.max_row if isinstance(ws.max_row, int) else 0
    n_cols = ws.max_column if isinstance(ws.max_column, int) else 0
    dates: list = []
    nums: list = []
    n_dates = n_nums = 0
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, bool):
                continue
            if isinstance(v, (_dt.datetime, _dt.date)):
                n_dates += 1
                d = v.date() if isinstance(v, _dt.datetime) else v
                if not dates:
                    dates = [d, d]
                else:
                    dates[0] = min(dates[0], d)
                    dates[1] = max(dates[1], d)
            elif isinstance(v, (int, float)):
                n_nums += 1
                if not nums:
                    nums = [v, v]
                else:
                    nums[0] = min(nums[0], v)
                    nums[1] = max(nums[1], v)
    bits = [f"{n_rows} rows x {n_cols} cols"]
    if n_dates:
        bits.append(f"{n_dates} date cells spanning {dates[0].isoformat()}..{dates[1].isoformat()}")
    if n_nums:
        bits.append(f"{n_nums} numeric cells in [{nums[0]:g}, {nums[1]:g}]")
    return "*(sheet facts: " + "; ".join(bits) + ")*"


class XlsxExtractor(Extractor):
    name = f"openpyxl@{_VERSION}"
    extensions = frozenset({".xlsx", ".xlsm"})

    def extract(self, raw: bytes) -> str:
        wb = _openpyxl.load_workbook(
            io.BytesIO(raw), read_only=True, data_only=True
        )
        parts: list[str] = []
        for ws in wb.worksheets:
            parts.append(f"## {ws.title}")
            parts.append("")
            parts.append(_sheet_facts(ws))
            parts.append("")
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                parts.append("*(empty sheet)*")
                parts.append("")
                continue
            cells = [_cell(v) for v in header]
            parts.append("| " + " | ".join(cells) + " |")
            parts.append("|" + "---|" * len(cells))
            emitted = 1
            truncated = False
            for row in rows:
                if emitted >= MAX_ROWS_PER_SHEET:
                    truncated = True
                    break
                parts.append("| " + " | ".join(_cell(v) for v in row) + " |")
                emitted += 1
            if truncated:
                # max_row comes from the sheet's declared dimensions; read-only
                # mode reports it without walking the remaining rows.
                total = ws.max_row if isinstance(ws.max_row, int) else None
                of = f" of {total}" if total else ""
                parts.append("")
                parts.append(
                    f"*(first {emitted}{of} rows — the canonical .xlsx holds "
                    "the rest)*"
                )
            parts.append("")
        wb.close()
        parts.append(
            "*(values as last calculated, not formulas — openpyxl data_only)*"
        )
        return "\n".join(parts)
